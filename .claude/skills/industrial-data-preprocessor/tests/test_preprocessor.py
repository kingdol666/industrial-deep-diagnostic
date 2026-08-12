#!/usr/bin/env python3
"""Unit tests for the adaptive data preprocessor (E-1).

Verifies the pipeline accepts ANY data source and yields the canonical
preprocessed_data.csv + report:

1.  single clean CSV            — passthrough
2.  GBK + semicolon + junk head — encoding/delimiter/header adaptation
3.  multi-sheet xlsx w/ junk    — main sheet selection
4.  xlsm workbook               — macro-enabled Excel support
5.  JSON record array           — json adaptation
6.  markdown tables / free text — table extraction vs context preservation
7.  mixed directory             — csv+xlsx+md+py → merge/largest + context
8.  batch data without time     — cross-sectional passthrough
9.  multi-block delimited file  — largest block wins
10. no-tabular directory        — graceful no_tabular_data report

Run: uv run --project .claude/shared/scripts python .claude/skills/industrial-data-preprocessor/tests/test_preprocessor.py
"""

from __future__ import annotations

import json
import shutil
import sys
import tempfile
from pathlib import Path

import numpy as np
import pandas as pd

_HERE = Path(__file__).resolve().parent
_SCRIPTS = _HERE.parent / "scripts"
sys.path.insert(0, str(_SCRIPTS))

from data_preprocessor import preprocess  # noqa: E402

_FAILURES: list[str] = []
_TMP: Path = Path(tempfile.mkdtemp(prefix="prep_test_"))


def check(name: str, cond: bool, detail: str = "") -> None:
    if not cond:
        _FAILURES.append(f"{name}: {detail}")
        print(f"  FAIL {name}: {detail}")
    else:
        print(f"  ok   {name}")


def run_case(data_path: Path, out_dir: Path, name: str = "t") -> dict:
    return preprocess(data_path, out_dir, name)


def test_clean_csv() -> None:
    d = _TMP / "t1"
    out = d / "out"
    d.mkdir(parents=True)
    pd.DataFrame({
        "timestamp": pd.date_range("2025-01-01", periods=100, freq="min"),
        "pressure_bar": np.random.normal(10, 1, 100),
        "temp_C": np.random.normal(200, 5, 100),
    }).to_csv(d / "data.csv", index=False)
    rep = run_case(d / "data.csv", out)
    df = pd.read_csv(out / "preprocessed_data.csv")
    check("t1-csv-shape", df.shape == (100, 3), str(df.shape))
    check("t1-time-col", rep["quality"]["time_column"] == "timestamp",
          str(rep["quality"]["time_column"]))


def test_gbk_semicolon_junk() -> None:
    d = _TMP / "t2"
    out = d / "out"
    d.mkdir(parents=True)
    lines = [
        "设备运行数据导出报告",              # junk title
        "导出时间：2025-01-02",              # junk
        "温度(℃);压力(MPa);流量(m³/h);时间",  # header with semicolons
    ]
    for i in range(120):
        lines.append(f"{210 + i * 0.1:.1f};{8 + i * 0.01:.3f};{55 + i * 0.05:.2f};2025-01-02 0{i // 60}:{i % 60:02d}:00")
    (d / "设备导出.txt").write_bytes("\n".join(lines).encode("gb18030"))
    rep = run_case(d / "设备导出.txt", out)
    df = pd.read_csv(out / "preprocessed_data.csv")
    check("t2-gbk-rows", df.shape[0] == 120, str(df.shape))
    check("t2-gbk-cols", df.shape[1] == 4, str(df.shape))
    check("t2-numeric", pd.to_numeric(df.iloc[0, 0], errors="coerce") == 210.0,
          str(df.iloc[0, 0]))


def test_multisheet_xlsx() -> None:
    d = _TMP / "t3"
    out = d / "out"
    d.mkdir(parents=True)
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["junk"])
    ws2 = wb.create_sheet("YAMATAKE_QUERY")
    ws2.append(["a"])
    ws3 = wb.create_sheet("数据统计")
    ws3.append(["timestamp", "value_1", "value_2"])
    for i in range(300):
        ws3.append([f"2025-01-01 00:{i % 60:02d}:00", 1.0 + i, 2.0 * i])
    wb.save(d / "multi.xlsx")
    rep = run_case(d / "multi.xlsx", out)
    df = pd.read_csv(out / "preprocessed_data.csv")
    check("t3-xlsx-shape", df.shape == (300, 3), str(df.shape))
    check("t3-sheet-selected", rep["selected_table"]["sheet"] == "数据统计",
          str(rep["selected_table"]))


def test_xlsm() -> None:
    d = _TMP / "t4"
    out = d / "out"
    d.mkdir(parents=True)
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "宏数据"
    ws.append(["time", "temp", "torque"])
    for i in range(200):
        ws.append([f"2025-02-01 {i // 60:02d}:{i % 60:02d}:00", 30 + i * 0.1, 50 - i * 0.05])
    wb.save(d / "macro_data.xlsm")
    rep = run_case(d / "macro_data.xlsm", out)
    df = pd.read_csv(out / "preprocessed_data.csv")
    check("t4-xlsm-shape", df.shape == (200, 3), str(df.shape))
    check("t4-xlsm-time", rep["quality"]["time_column"] == "time")


def test_json_records() -> None:
    d = _TMP / "t5"
    out = d / "out"
    d.mkdir(parents=True)
    rows = [{"ts": f"2025-03-01 00:{i // 60:02d}:{i % 60:02d}", "speed": 1000 + i, "load": i * 0.5}
            for i in range(150)]
    (d / "data.json").write_text(json.dumps(rows), encoding="utf-8")
    rep = run_case(d / "data.json", out)
    df = pd.read_csv(out / "preprocessed_data.csv")
    check("t5-json-shape", df.shape == (150, 3), str(df.shape))
    check("t5-json-time", rep["quality"]["time_column"] == "ts")


def test_markdown() -> None:
    d = _TMP / "t6"
    out = d / "out"
    d.mkdir(parents=True)
    md = """# 工艺说明

本文件描述流程背景。

| 时间 | 温度 | 压力 |
|------|------|------|
| 2025-01-01 00:00:00 | 200 | 8 |
| 2025-01-01 00:01:00 | 201 | 8.1 |
| 2025-01-01 00:02:00 | 202 | 8.2 |
| 2025-01-01 00:03:00 | 203 | 8.3 |
| 2025-01-01 00:04:00 | 204 | 8.4 |
| 2025-01-01 00:05:00 | 205 | 8.5 |
"""
    (d / "process_table.md").write_text(md, encoding="utf-8")
    rep = run_case(d / "process_table.md", out)
    df = pd.read_csv(out / "preprocessed_data.csv")
    check("t6-md-shape", df.shape == (6, 3), str(df.shape))

    # free-text md → context, no_tabular
    out2 = d / "out2"
    (d / "notes.md").write_text("# 说明\n没有表格的自由文本。", encoding="utf-8")
    rep2 = run_case(d / "notes.md", out2)
    check("t6-md-context", rep2["status"] == "no_tabular_data"
          and len(rep2["context_files"]) == 1, str(rep2["status"]))

    # tiny md table (<5 rows) → context, not data
    out3 = d / "out3"
    (d / "snippet.md").write_text("| a | b |\n|---|---|\n| 1 | 2 |\n| 3 | 4 |\n", encoding="utf-8")
    rep3 = run_case(d / "snippet.md", out3)
    check("t6-md-tiny-context", rep3["status"] == "no_tabular_data"
          and len(rep3["context_files"]) == 1, str(rep3["status"]))


def test_mixed_directory() -> None:
    d = _TMP / "t7"
    out = d / "out"
    d.mkdir(parents=True)
    # process csv
    pd.DataFrame({
        "timestamp": pd.date_range("2025-01-01", periods=200, freq="min"),
        "pressure": np.random.normal(8, 0.5, 200),
    }).to_csv(d / "process.csv", index=False)
    # quality xlsx
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "quality"
    ws.append(["timestamp", "quality_score"])
    for i in range(200):
        ws.append([f"2025-01-01 {i // 60:02d}:{i % 60:02d}:00", 90 + i * 0.02])
    wb.save(d / "quality.xlsx")
    # context + skipped
    (d / "ground_truth.md").write_text("# 真值说明", encoding="utf-8")
    (d / "gen.py").write_text("print('x')", encoding="utf-8")
    rep = run_case(d, out)
    df = pd.read_csv(out / "preprocessed_data.csv")
    check("t7-merge", df.shape[0] >= 200 and "pressure" in df.columns
          and "quality_score" in df.columns, str(df.shape) + str(list(df.columns)))
    check("t7-context", any("ground_truth.md" in f for f in rep["context_files"]), str(rep["context_files"]))
    check("t7-skip", any(s["file"] == "gen.py" for s in rep["skipped_files"]), str(rep["skipped_files"]))


def test_batch_no_time() -> None:
    d = _TMP / "t8"
    out = d / "out"
    d.mkdir(parents=True)
    pd.DataFrame({
        "batch_id": [f"B{i}" for i in range(50)],
        "yield_pct": np.random.normal(85, 3, 50),
        "impurity_ppm": np.random.normal(20, 5, 50),
    }).to_csv(d / "batch.csv", index=False)
    rep = run_case(d / "batch.csv", out)
    df = pd.read_csv(out / "preprocessed_data.csv")
    check("t8-batch-shape", df.shape == (50, 3), str(df.shape))
    check("t8-no-time", rep["quality"]["time_column"] is None)


def test_multiblock_delimited() -> None:
    d = _TMP / "t9"
    out = d / "out"
    d.mkdir(parents=True)
    big = ["a,b,c"] + [f"1,2,3"] * 400
    small = ["x,y"] + ["9,9"] * 10
    (d / "blocks.txt").write_text("\n".join(big + [""] + small), encoding="utf-8")
    rep = run_case(d / "blocks.txt", out)
    df = pd.read_csv(out / "preprocessed_data.csv")
    check("t9-block-largest", df.shape == (400, 3), str(df.shape))


def test_no_tabular_dir() -> None:
    d = _TMP / "t10"
    out = d / "out"
    d.mkdir(parents=True)
    (d / "notes.md").write_text("# 只有说明\n没有表格。", encoding="utf-8")
    (d / "plan.py").write_text("pass", encoding="utf-8")
    rep = run_case(d, out)
    check("t10-status", rep["status"] == "no_tabular_data", rep["status"])
    check("t10-context-kept", len(rep["context_files"]) == 1, str(rep["context_files"]))


def main() -> None:
    print("preprocessor tests:")
    for fn in [
        test_clean_csv,
        test_gbk_semicolon_junk,
        test_multisheet_xlsx,
        test_xlsm,
        test_json_records,
        test_markdown,
        test_mixed_directory,
        test_batch_no_time,
        test_multiblock_delimited,
        test_no_tabular_dir,
    ]:
        fn()
    shutil.rmtree(_TMP, ignore_errors=True)
    if _FAILURES:
        print(f"\n{len(_FAILURES)} FAILURES")
        sys.exit(1)
    print("\nALL TESTS PASSED")


if __name__ == "__main__":
    main()
