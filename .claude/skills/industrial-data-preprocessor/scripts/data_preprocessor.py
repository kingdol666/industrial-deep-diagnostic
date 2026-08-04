#!/usr/bin/env python3
"""data_preprocessor.py — E-1: adaptive data preprocessing stage.

Turns ANY user-supplied data source — a single file or a whole directory, in
any mix of CSV/TSV/delimited text, Excel (.xlsx/.xlsm/.xls), JSON records,
Markdown tables, HTML tables, or free-text notes — into ONE canonical table
that the industrial diagnostic pipeline can consume unchanged.

Deterministic. Depends only on the Python standard library + pandas +
openpyxl (already required by the pipeline). Never modifies the input.

Outputs (written under ``--output``, typically ``RUN_DIR/00_input/``):

* ``preprocessed_data.csv``      — canonical tabular data (the pipeline input)
* ``preprocessed_data.json``     — same data as JSON records (fallback reader)
* ``preprocessing_report.json``  — full audit: per-source disposition, format
                                   detection, merge keys, quality flags
* ``context/``                   — non-tabular files (process descriptions,
                                   ground truth, notes) preserved verbatim for
                                   the ontology/context builder

CLI::

    python data_preprocessor.py --data-path <file|dir> --output <dir>
                                [--name <run_name>]

Exit 0 with a report even when nothing tabular is found (the report then says
``no_tabular_data`` and the caller decides); exit 1 only on hard errors.
"""

from __future__ import annotations

import argparse
import json
import re
import shutil
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd

# ---------------------------------------------------------------------------
# Table extraction: one source file → list of candidate tables
# ---------------------------------------------------------------------------

_CONTEXT_EXTS = {".md", ".markdown", ".txt", ".rst", ".doc", ".docx", ".pdf"}
_SKIP_EXTS = {
    ".py", ".mjs", ".js", ".ts", ".ipynb", ".sh", ".bat", ".db", ".db-shm",
    ".db-wal", ".png", ".jpg", ".jpeg", ".gif", ".svg", ".pdf", ".docx",
    ".doc", ".zip", ".tar", ".gz", ".7z", ".gitkeep", ".ini", ".cfg", ".yml",
    ".yaml", ".log", ".bak",
}
_TABLE_EXTS = {".csv", ".tsv", ".txt", ".dat", ".xlsx", ".xlsm", ".xls",
               ".json", ".md", ".markdown", ".html", ".htm"}
_TIME_HINTS = ("time", "timestamp", "datetime", "date", "zeit", "ts_", "时间",
               "日期", "时刻")
_MAIN_SHEET_HINTS = ("data", "记录", "统计", "历史", "history", "export", "查询")
_JUNK_SHEET_HINTS = ("query", "sheet", "说明", "注释", "note", "readme")


def _read_text_adaptive(path: Path) -> Tuple[str, str]:
    """Read text with encoding fallback: utf-8(-sig) → gb18030 → latin-1."""
    raw = path.read_bytes()
    for enc in ("utf-8-sig", "utf-8", "gb18030", "latin-1"):
        try:
            return raw.decode(enc), enc
        except (UnicodeDecodeError, LookupError):
            continue
    return raw.decode("utf-8", errors="replace"), "utf-8(replace)"


def _detect_delimiter(lines: List[str]) -> str:
    """Delimiter detection over the first non-empty lines: the delimiter that
    appears in the most lines wins (junk title lines must not veto the real
    delimiter present in header/data rows)."""
    sample = [ln.strip() for ln in lines[:12] if ln.strip()]
    if not sample:
        return ","
    cands = {"\t": 0, ";": 0, ",": 0, "|": 0}
    for ln in sample:
        for d in cands:
            if d in ln:
                cands[d] += 1
    best = max(cands, key=lambda d: (cands[d], d != ","))
    # require the winner to actually appear somewhere
    if cands[best] == 0:
        return ","
    return best


def _split_line(line: str, delim: str) -> List[str]:
    if delim == "\t":
        return [c.strip() for c in line.rstrip("\n").split("\t")]
    if delim == ";":
        return [c.strip() for c in line.rstrip("\n").split(";")]
    if delim == "|":
        return [c.strip() for c in line.rstrip("\n").split("|")]
    # comma with optional quotes
    out: List[str] = []
    cur: List[str] = []
    in_q = False
    for ch in line.rstrip("\n"):
        if ch == '"':
            in_q = not in_q
        elif ch == "," and not in_q:
            out.append("".join(cur).strip())
            cur = []
        else:
            cur.append(ch)
    out.append("".join(cur).strip())
    return out


def _is_header_row(cells: List[str]) -> bool:
    """Header heuristics: mostly non-numeric, short, unique."""
    if not cells:
        return False
    nonempty = [c for c in cells if c != ""]
    if not nonempty:
        return False
    numeric = 0
    for c in nonempty:
        try:
            float(c.replace(",", "").replace("%", "").replace(" ", ""))
            numeric += 1
        except ValueError:
            pass
    return numeric / len(nonempty) < 0.5


def _numeric_density(cells: List[str]) -> float:
    if not cells:
        return 0.0
    ok = 0
    for c in cells:
        if c == "" or c is None:
            continue
        try:
            float(str(c).replace(",", "").replace("%", "").strip())
            ok += 1
        except ValueError:
            pass
    return ok / max(1, len([c for c in cells if c not in ("", None)]))


def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Trim names, de-dup, drop fully-empty columns."""
    cols: List[str] = []
    seen: Dict[str, int] = {}
    for c in df.columns:
        name = str(c).strip()
        if not name:
            name = "column"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        cols.append(name)
    df.columns = cols
    blank = df.apply(lambda s: s.astype(str).str.strip() == "")
    df = df.loc[:, ~(df.isna().all() | blank.all())]
    return df.reset_index(drop=True)


def _parse_delimited_text(path: Path) -> List[dict]:
    """Parse CSV/TSV/delimited text, skipping junk header lines; returns
    candidate tables (multi-block files yield several candidates)."""
    text, enc = _read_text_adaptive(path)
    lines = text.splitlines()
    if not lines:
        return []
    delim = _detect_delimiter(lines)
    # find the first plausible header row (skip title/junk lines)
    start = 0
    for i, ln in enumerate(lines[:20]):
        cells = _split_line(ln, delim)
        if len(cells) >= 2 and _is_header_row(cells):
            start = i
            break
    blocks: List[dict] = []
    cur_rows: List[List[str]] = []
    header: Optional[List[str]] = None
    n_cols = 0

    def flush() -> None:
        nonlocal cur_rows, header, n_cols
        if header is not None and len(cur_rows) >= 2:
            blocks.append({"source": path.name, "sheet": None,
                           "format": "delimited", "encoding": enc,
                           "header": header, "rows": cur_rows})
        cur_rows = []
        header = None

    for i, ln in enumerate(lines[start:], start=start):
        if not ln.strip():
            flush()
            continue
        cells = _split_line(ln, delim)
        if header is None:
            if len(cells) >= 2 and _is_header_row(cells):
                header = cells
                n_cols = len(cells)
            continue
        # start of a new block: a new header-like line with different width
        if (len(cells) >= 2 and _is_header_row(cells)
                and len(cells) != n_cols and len(cur_rows) >= 2):
            flush()
            header = cells
            n_cols = len(cells)
            continue
        # pad/trim to header width
        while len(cells) < n_cols:
            cells.append("")
        cur_rows.append(cells[:n_cols])
    flush()
    return blocks


def _sheet_score(sheet_name: str, n_rows: int, n_cols: int,
                 numeric_density: float) -> float:
    """Score a worksheet: big real-data sheets win; junk sheets lose."""
    score = 0.0
    if n_rows >= 100:
        score += 4.0
    elif n_rows >= 10:
        score += 2.0
    score += min(n_cols / 10.0, 2.0)
    score += numeric_density * 2.0
    low = sheet_name.lower()
    if any(k in low for k in _JUNK_SHEET_HINTS):
        score -= 5.0
    if any(k in low for k in _MAIN_SHEET_HINTS):
        score += 1.5
    if n_rows <= 2 or n_cols <= 1:
        score -= 10.0
    return score


def _parse_excel(path: Path) -> List[dict]:
    """Parse .xlsx/.xlsm/.xls: every sheet becomes a candidate; junk sheets
    (1x1, 1-row query stubs) are scored down but preserved in the report."""
    try:
        import openpyxl
    except ImportError:
        return [{"source": path.name, "error": "openpyxl not installed",
                 "format": "excel"}]
    blocks: List[dict] = []
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # pragma: no cover
        return [{"source": path.name, "error": f"excel open failed: {exc}",
                 "format": "excel"}]
    try:
        for ws in wb.worksheets:
            n_rows = ws.max_row or 0
            n_cols = ws.max_column or 0
            if n_rows <= 1:
                continue
            # stream rows (read_only mode)
            rows: List[List[str]] = []
            for r in ws.iter_rows(values_only=True):
                cells = ["" if v is None else str(v).strip() for v in r]
                rows.append(cells)
            if len(rows) < 2:
                continue
            # find header: first row that is not junk
            header_idx = 0
            for i, r in enumerate(rows[:5]):
                if _is_header_row(r) and len([c for c in r if c]) >= 1:
                    header_idx = i
                    break
            header = rows[header_idx]
            data_rows = rows[header_idx + 1:]
            n_cols = len(header)
            data_rows = [r[:n_cols] + [""] * (n_cols - len(r)) for r in data_rows]
            # numeric density of first 200 data rows
            sample = [c for r in data_rows[:200] for c in r]
            density = _numeric_density(sample)
            blocks.append({
                "source": path.name, "sheet": ws.title,
                "format": "excel", "encoding": "xlsx",
                "header": header, "rows": data_rows,
                "score": _sheet_score(ws.title, len(data_rows), n_cols, density),
                "n_rows": len(data_rows), "n_cols": n_cols,
                "numeric_density": round(density, 3),
            })
    finally:
        wb.close()
    return blocks


def _parse_json(path: Path) -> List[dict]:
    """Parse JSON: array of records, {data: [...]}, or single object."""
    text, enc = _read_text_adaptive(path)
    try:
        obj = json.loads(text)
    except json.JSONDecodeError as exc:
        return [{"source": path.name, "error": f"json parse failed: {exc}",
                 "format": "json"}]
    arr = None
    if isinstance(obj, list):
        arr = obj
    elif isinstance(obj, dict):
        for key in ("data", "records", "rows", "values", "items", "result"):
            if isinstance(obj.get(key), list) and obj[key]:
                arr = obj[key]
                break
        if arr is None:
            # single record object
            arr = [obj]
    if not arr:
        return [{"source": path.name, "error": "empty json", "format": "json"}]
    if not all(isinstance(r, dict) for r in arr):
        return [{"source": path.name,
                 "error": "json is not a record array", "format": "json"}]
    df = pd.DataFrame(arr)
    df = _normalize_columns(df)
    blocks = [{
        "source": path.name, "sheet": None, "format": "json", "encoding": enc,
        "header": list(df.columns), "rows": df.astype(str).values.tolist(),
    }]
    return blocks


_MD_TABLE_RE = re.compile(
    r"^\s*\|(.+)\|\s*$", re.MULTILINE)


def _parse_markdown(path: Path) -> List[dict]:
    """Extract pipe tables from Markdown; if none, the file is CONTEXT."""
    text, _ = _read_text_adaptive(path)
    lines = text.splitlines()
    tables: List[dict] = []
    cur: List[List[str]] = []
    in_table = False
    for ln in lines:
        if _MD_TABLE_RE.match(ln):
            cells = [c.strip() for c in ln.strip().strip("|").split("|")]
            if not in_table:
                cur = []
                in_table = True
            # skip separator rows like |---|---|
            if all(re.fullmatch(r":?-{2,}:?", c) for c in cells if c.strip()):
                continue
            cur.append(cells)
        else:
            if in_table and len(cur) >= 2:
                tables.append({"source": path.name, "sheet": None,
                               "format": "markdown", "encoding": "utf-8",
                               "header": cur[0], "rows": cur[1:]})
            in_table = False
    if in_table and len(cur) >= 2:
        tables.append({"source": path.name, "sheet": None,
                       "format": "markdown", "encoding": "utf-8",
                       "header": cur[0], "rows": cur[1:]})
    return tables


_HTML_TABLE_RE = re.compile(r"<table[^>]*>(.*?)</table>", re.IGNORECASE | re.DOTALL)
_HTML_ROW_RE = re.compile(r"<tr[^>]*>(.*?)</tr>", re.IGNORECASE | re.DOTALL)
_HTML_CELL_RE = re.compile(r"<t[dh][^>]*>(.*?)</t[dh]>", re.IGNORECASE | re.DOTALL)


def _strip_tags(html: str) -> str:
    return re.sub(r"<[^>]+>", "", html).replace("&nbsp;", " ").strip()


def _parse_html(path: Path) -> List[dict]:
    """Lightweight regex-based HTML table extractor (no lxml dependency)."""
    text, _ = _read_text_adaptive(path)
    tables: List[dict] = []
    for m in _HTML_TABLE_RE.finditer(text):
        rows = []
        for rm in _HTML_ROW_RE.finditer(m.group(1)):
            cells = [_strip_tags(cm.group(1)) for cm in _HTML_CELL_RE.finditer(rm.group(1))]
            if cells:
                rows.append(cells)
        if len(rows) >= 2:
            tables.append({"source": path.name, "sheet": None,
                           "format": "html", "encoding": "utf-8",
                           "header": rows[0], "rows": rows[1:]})
    return tables


# ---------------------------------------------------------------------------
# Candidate → DataFrame
# ---------------------------------------------------------------------------

def _block_to_df(block: dict) -> pd.DataFrame:
    df = pd.DataFrame(block["rows"], columns=block["header"])
    df = _normalize_columns(df)
    # YAMATAKE-style exports repeat the header row inside the data stream —
    # drop any row that is byte-identical to the (normalized) header.
    if df.shape[0] > 0 and df.shape[1] > 0:
        hdr = np.array([str(h) for h in df.columns])
        if len(hdr) == df.shape[1]:
            mask = (df.astype(str).to_numpy() == hdr).all(axis=1)
            if mask.any():
                df = df.loc[~mask].reset_index(drop=True)
    return df


def _normalize_cn_datetime(value: str) -> str:
    """Normalize Chinese calendar formats pandas cannot parse natively:
    '2026年5月1日 08:00:00' -> '2026-05-01 08:00:00'. Also handles
    '2026/5/1 8:00' and '2026.5.1 08:00' variants."""
    s = str(value).strip()
    m = re.match(
        r"(\d{4})\s*[年./-]\s*(\d{1,2})\s*[月./-]\s*(\d{1,2})\s*[日]?"
        r"(?:\s+(\d{1,2}):(\d{2})(?::(\d{2}))?)?", s)
    if m:
        y, mo, d = m.group(1), int(m.group(2)), int(m.group(3))
        hh, mm, ss = m.group(4) or "0", m.group(5) or "0", m.group(6) or "00"
        return f"{y}-{mo:02d}-{d:02d} {hh}:{mm}:{ss}"
    return s


def _parses_as_time(values) -> float:
    """Fraction of values parseable as datetime after CN-format normalization."""
    normed = [_normalize_cn_datetime(v) for v in values]
    try:
        parsed = pd.to_datetime(pd.Series(normed), errors="coerce")
        return float(parsed.notna().mean())
    except Exception:
        return 0.0


def _detect_time_col(df: pd.DataFrame) -> Optional[str]:
    for c in df.columns:
        low = str(c).lower()
        if any(k in low for k in _TIME_HINTS):
            # verify it parses as datetime (at least partially)
            sample = df[c].dropna().head(50).astype(str)
            if _parses_as_time(sample) > 0.5:
                return str(c)
    for c in df.columns:
        # numeric columns would parse as nanosecond-epoch datetimes — skip them
        try:
            if pd.to_numeric(df[c], errors="coerce").notna().mean() > 0.7:
                continue
        except Exception:
            continue
        if _parses_as_time(df[c].dropna().head(100).astype(str)) > 0.8:
            return str(c)
    return None


def _looks_premerged(name: str) -> bool:
    """A file that already IS the merged/alignment product of the raw exports
    (e.g. merged_process_data_full.csv) must be used as-is, never re-merged."""
    low = Path(name).stem.lower()
    return any(k in low for k in ("merged", "merge", "align", "aligned",
                                  "combined", "汇总", "合并", "对齐", "_full",
                                  "joined"))


def _score_table(df: pd.DataFrame, time_col: Optional[str]) -> float:
    n_rows, n_cols = df.shape
    if n_cols == 0:
        return -1.0
    numeric = 0
    for c in df.columns:
        try:
            if pd.to_numeric(df[c], errors="coerce").notna().mean() > 0.7:
                numeric += 1
        except Exception:
            pass
    score = min(n_rows / 1000.0, 3.0) + min(n_cols / 5.0, 3.0) + numeric * 0.5
    if time_col:
        score += 1.0
    return score


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------

def _discover_sources(data_path: Path) -> List[Path]:
    if data_path.is_file():
        return [data_path]
    out: List[Path] = []
    for p in sorted(data_path.iterdir()):
        if p.is_dir():
            continue
        out.append(p)
    return out


def _classify(path: Path) -> str:
    ext = path.suffix.lower()
    if ext in _TABLE_EXTS:
        return "table"
    if ext in _CONTEXT_EXTS:
        return "context"
    return "skip"


def preprocess(data_path: Path, output_dir: Path, run_name: str = "") -> dict:
    """Main entry: discover → extract → select/merge → normalize → emit."""
    output_dir.mkdir(parents=True, exist_ok=True)
    context_dir = output_dir / "context"
    context_dir.mkdir(parents=True, exist_ok=True)

    sources = _discover_sources(data_path)
    if not sources:
        return _fail_report(output_dir, run_name, "no files found", str(data_path))

    report: Dict[str, Any] = {
        "run_name": run_name,
        "input_path": str(data_path),
        "input_is_directory": data_path.is_dir(),
        "sources": [],
        "tables": [],
        "selected_table": None,
        "merge": None,
        "quality": {},
        "context_files": [],
        "skipped_files": [],
        "status": "ok",
    }

    candidates: List[Tuple[str, dict]] = []  # (file_name, block)
    for src in sources:
        cls = _classify(src)
        if cls == "skip":
            report["skipped_files"].append({"file": src.name,
                                            "reason": f"extension {src.suffix or '(none)'} not tabular/context"})
            continue
        if cls == "context":
            dest = context_dir / src.name
            shutil.copy2(src, dest)
            report["context_files"].append(str(dest))
            report["sources"].append({"file": src.name, "role": "context",
                                      "reason": "non-tabular document preserved for ontology/context builder"})
            continue

        ext = src.suffix.lower()
        blocks: List[dict] = []
        if ext in (".csv", ".tsv", ".txt", ".dat"):
            blocks = _parse_delimited_text(src)
        elif ext in (".xlsx", ".xlsm", ".xls"):
            blocks = _parse_excel(src)
        elif ext == ".json":
            blocks = _parse_json(src)
        elif ext in (".md", ".markdown"):
            blocks = _parse_markdown(src)
        elif ext in (".html", ".htm"):
            blocks = _parse_html(src)

        if not blocks:
            report["sources"].append({"file": src.name, "role": "no_tabular",
                                      "reason": "no tabular content detected"})
            if ext in (".md", ".markdown", ".txt"):
                dest = context_dir / src.name
                shutil.copy2(src, dest)
                report["context_files"].append(str(dest))
            continue

        if any("error" in b for b in blocks):
            for b in blocks:
                if "error" in b:
                    report["sources"].append({"file": src.name, "role": "error",
                                              "reason": b["error"]})
            continue

        # materialize every block as a scored DataFrame candidate
        for bi, b in enumerate(blocks):
            df = _block_to_df(b)
            if df.shape[0] < 2 or df.shape[1] < 1:
                continue
            # Markdown/HTML tables with <5 data rows are snippets (e.g. a table
            # inside a process description), not the dataset — keep as context.
            if b.get("format") in ("markdown", "html") and df.shape[0] < 5:
                report["sources"].append({"file": src.name, "role": "context",
                                          "reason": f"tiny {b.get('format')} table ({df.shape[0]} rows) — document snippet"})
                dest = context_dir / src.name
                if not dest.exists():
                    shutil.copy2(src, dest)
                    report["context_files"].append(str(dest))
                continue
            tcol = _detect_time_col(df)
            score = _score_table(df, tcol)
            cand = {
                "file": src.name,
                "sheet": b.get("sheet"),
                "format": b.get("format", ext.lstrip(".")),
                "n_rows": int(df.shape[0]),
                "n_cols": int(df.shape[1]),
                "time_column": tcol,
                "score": round(score, 3),
                "numeric_density": round(b.get("numeric_density", _numeric_density(
                    [str(v) for v in df.iloc[:200].to_numpy().ravel()])), 3),
                "block_index": bi,
                "df": df,
            }
            candidates.append((src.name, cand))
            report["tables"].append({
                "file": src.name,
                "sheet": b.get("sheet"),
                "format": b.get("format"),
                "n_rows": int(df.shape[0]),
                "n_cols": int(df.shape[1]),
                "time_column": tcol,
                "score": round(score, 3),
            })

    if not candidates:
        report["status"] = "no_tabular_data"
        _write_report(output_dir, report)
        return report

    # ---- selection / merge --------------------------------------------
    primary: Optional[Tuple[str, dict]] = None
    merge_info: Optional[dict] = None
    selected_kind = "score"

    # 1) already-merged product files win outright (user-aligned final data)
    premerged = [c for c in candidates
                 if _looks_premerged(c[0]) and c[1]["n_rows"] >= 100]
    if premerged:
        primary = max(premerged, key=lambda c: c[1]["n_rows"])
        selected_kind = "premerged_file"
    if primary is None:
        # 2) Excel sheet whose name matches the file stem is THE data
        for fname, cand in candidates:
            stem = Path(fname).stem.lower()
            if cand["sheet"] and cand["sheet"].lower() == stem and cand["n_rows"] >= 100:
                primary = (fname, cand)
                selected_kind = "excel_stem_match"
                break
    if primary is None:
        # 3) largest by score
        primary = max(candidates, key=lambda c: c[1]["score"])

    # 3) attempt merge ONLY when no premerged product exists: several raw
    #    exports sharing a time axis, bounded by an anti-bloat gate so a failed
    #    join can never produce a garbage mega-table.
    large = [c for c in candidates if c[1]["n_rows"] >= 100]
    if (selected_kind != "premerged_file" and len(large) >= 2
            and primary is not None and primary[1]["time_column"]):
        pname, pcand = primary
        ptime = pcand["time_column"]

        def _time_axis(cand: dict) -> Optional[np.ndarray]:
            tc = cand["time_column"]
            if not tc:
                return None
            try:
                parsed = pd.to_datetime(cand["df"][tc], errors="coerce")
                return parsed.astype("int64").to_numpy()
            except Exception:
                return None

        base_axis = _time_axis(pcand)
        merged_dfs: List[Tuple[str, pd.DataFrame, str]] = []
        for fname, cand in large:
            ctime = cand["time_column"]
            if not (ctime and base_axis is not None):
                continue
            axis = _time_axis(cand)
            if axis is None or len(axis) == 0:
                continue
            # overlap = |intersection of ranges| / |smaller range|
            lo = max(float(np.nanmin(axis)), float(np.nanmin(base_axis)))
            hi = min(float(np.nanmax(axis)), float(np.nanmax(base_axis)))
            span = min(float(np.nanmax(axis)) - float(np.nanmin(axis)),
                       float(np.nanmax(base_axis)) - float(np.nanmin(base_axis)))
            if span <= 0 or (hi - lo) / span < 0.3:
                continue
            merged_dfs.append((fname, cand["df"], ctime))
        if len(merged_dfs) >= 2:
            merge_info = {
                "strategy": "outer_merge_on_time_column",
                "key": ptime,
                "note": "sheets matched by time-axis overlap >=30%; values normalized to ISO before join",
                "merged_files": sorted({f for f, _, _ in merged_dfs}),
            }
            base = merged_dfs[0][1].copy()
            base_time = merged_dfs[0][2]
            base[base_time] = pd.to_datetime(base[base_time], errors="coerce").dt.strftime("%Y-%m-%d %H:%M:%S")
            for fname, df, ctime in merged_dfs[1:]:
                right = df.copy()
                right[ctime] = pd.to_datetime(right[ctime], errors="coerce").dt.strftime("%Y-%m-%d %H:%M:%S")
                if ctime != base_time:
                    right = right.rename(columns={ctime: base_time})
                dup = [c for c in right.columns
                       if c in base.columns and c != base_time]
                right = right.drop(columns=dup)
                base = base.merge(right, on=base_time, how="outer", sort=True)
            base = _normalize_columns(base)
            # Anti-bloat gate: a clean time-axis join never grows beyond the
            # largest source table by much (rows = union of timestamps). If it
            # balloons (failed key alignment), abandon the merge and keep the
            # largest single table instead.
            max_src = max(m.shape[0] for _, m, _ in merged_dfs)
            if base.shape[0] > 1.5 * max_src:
                merge_info = {
                    "strategy": "abandoned",
                    "reason": f"merged rows {base.shape[0]} > 1.5x largest source ({max_src}); "
                              "time keys did not align — keeping largest single table",
                    "merged_files": sorted({f for f, _, _ in merged_dfs}),
                }
                primary = max(candidates, key=lambda c: c[1]["score"])
                selected_kind = "score_fallback"
            else:
                primary = (primary[0], {**primary[1], "df": base,
                                        "n_rows": int(base.shape[0]),
                                        "n_cols": int(base.shape[1]),
                                        "time_column": base_time})

    fname, cand = primary
    df = cand["df"].copy()
    report["selected_table"] = {
        "file": fname,
        "sheet": cand["sheet"],
        "format": cand["format"],
        "n_rows": int(df.shape[0]),
        "n_cols": int(df.shape[1]),
        "time_column": cand["time_column"],
        "score": cand["score"],
        "why": {
            "premerged_file": "file is an already-merged/alignment product of the raw exports",
            "excel_stem_match": "sheet name matches file stem",
            "score": "highest table score",
            "score_fallback": "merge abandoned (anti-bloat); largest single table",
        }.get(selected_kind, selected_kind),
    }
    if merge_info:
        report["merge"] = merge_info

    # ---- normalize + emit ---------------------------------------------
    time_col = cand["time_column"]
    if time_col is not None:
        try:
            parsed = pd.to_datetime(df[time_col], errors="coerce")
            # keep original values but report parseability; normalize
            # ISO-format copies only for clearly-parseable rows
            df[time_col] = parsed.dt.strftime("%Y-%m-%d %H:%M:%S").where(
                parsed.notna(), df[time_col].astype(str))
        except Exception:
            pass

    missing_pct = round(float(df.isna().mean().mean()) * 100, 2)
    report["quality"] = {
        "missing_cell_pct": missing_pct,
        "rows": int(df.shape[0]),
        "cols": int(df.shape[1]),
        "time_column": time_col,
        "numeric_columns": _numeric_cols(df),
        "note": "missing cells stay NaN; downstream cleaning handles them",
    }

    csv_out = output_dir / "preprocessed_data.csv"
    json_out = output_dir / "preprocessed_data.json"
    df.to_csv(csv_out, index=False, encoding="utf-8-sig")
    df.to_json(json_out, orient="records", force_ascii=False, indent=1)

    report["outputs"] = {
        "preprocessed_data.csv": str(csv_out),
        "preprocessed_data.json": str(json_out),
        "context_dir": str(context_dir),
    }
    report["status"] = "ok"
    _write_report(output_dir, report)
    return report


def _numeric_cols(df: pd.DataFrame) -> List[str]:
    out = []
    for c in df.columns:
        try:
            if pd.to_numeric(df[c], errors="coerce").notna().mean() > 0.7:
                out.append(str(c))
        except Exception:
            pass
    return out


def _fail_report(output_dir: Path, run_name: str, reason: str,
                 input_path: str) -> dict:
    report = {
        "run_name": run_name,
        "input_path": input_path,
        "status": "no_tabular_data",
        "reason": reason,
        "sources": [],
        "tables": [],
        "selected_table": None,
        "merge": None,
        "quality": {},
        "context_files": [],
        "skipped_files": [],
    }
    _write_report(output_dir, report)
    return report


def _write_report(output_dir: Path, report: dict) -> None:
    # strip live DataFrames before serializing
    clean = json.loads(json.dumps(report, default=lambda o: None, ensure_ascii=False))
    (output_dir / "preprocessing_report.json").write_text(
        json.dumps(clean, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def main() -> None:
    ap = argparse.ArgumentParser(description="E-1: adaptive data preprocessing")
    ap.add_argument("--data-path", required=True, help="file or directory")
    ap.add_argument("--output", required=True, help="output dir (RUN_DIR/00_input)")
    ap.add_argument("--name", default="", help="run name (optional)")
    args = ap.parse_args()

    data_path = Path(args.data_path)
    if not data_path.exists():
        print(f"ERROR: data path does not exist: {data_path}", file=sys.stderr)
        sys.exit(1)
    output = Path(args.output)
    report = preprocess(data_path, output, args.name)
    print(json.dumps({
        "status": report["status"],
        "selected": report.get("selected_table"),
        "tables_found": len(report.get("tables", [])),
        "context_files": len(report.get("context_files", [])),
        "skipped_files": len(report.get("skipped_files", [])),
        "report": str(output / "preprocessing_report.json"),
    }, indent=2, ensure_ascii=False))
    sys.exit(0 if report["status"] in ("ok", "no_tabular_data") else 1)


if __name__ == "__main__":
    main()
